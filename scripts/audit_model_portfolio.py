#!/usr/bin/env python3
"""
Model-portfolio integration auditor.

Reads the spreadsheet at:
  attached_assets/HakaseAI_PreTrained_Models_Portfolio_*.xlsx (latest)

…and grades every row in the "Master Portfolio" sheet against the actual
codebase, producing a wired/scaffolded/not-wired audit.

Outputs:
  artifacts/hakase-ai/docs/model-portfolio-audit.md
  artifacts/hakase-ai/docs/model-portfolio-audit.csv

Run:
  python3 scripts/audit_model_portfolio.py
  python3 scripts/audit_model_portfolio.py --xlsx <path>

The "evidence" for each model is grepped from a small, explicit set of
trees:
  - artifacts/ai-service/main.py + artifacts/ai-service/models/*.py
  - artifacts/hakase-ai/src/**/*.{ts,tsx}

Each portfolio model has an alias entry (a list of case-insensitive
regex patterns). If any alias hits a backend file we mark "live"; if it
hits a frontend lib but not a backend endpoint we mark "scaffolded";
otherwise "not wired".

Adding a new model? Add it to ALIASES below. The cost of a missing
entry is a "not wired" verdict, never a silent pass.
"""

from __future__ import annotations

import argparse
import csv
import glob
import os
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterable

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    sys.stderr.write("openpyxl is required: pip install openpyxl\n")
    sys.exit(2)


REPO_ROOT = Path(__file__).resolve().parent.parent
AI_SERVICE_DIR = REPO_ROOT / "artifacts" / "ai-service"
FRONTEND_SRC = REPO_ROOT / "artifacts" / "hakase-ai" / "src"
DOCS_DIR = REPO_ROOT / "artifacts" / "hakase-ai" / "docs"


# Each entry maps a portfolio "Model" cell value (lower-cased) to a list
# of regex patterns (case-insensitive) that, if present in a scanned
# file, count as evidence the model is integrated.
#
# Keep the patterns conservative — false positives here will quietly
# downgrade the audit's value. When in doubt, leave it out and let the
# verdict be "not wired" until someone wires it.
ALIASES: dict[str, list[str]] = {
    # === Foundation embeddings ===
    "chemberta-2": [r"chemberta", r"ChemBERTa-77M-MLM"],
    "molformer": [r"\bmolformer\b"],
    "uni-mol": [r"\buni-?mol\b"],
    "molbert": [r"\bmolbert\b"],
    "molencoder": [r"\bmolencoder\b", r"MolBART"],
    "naturelm": [r"\bnaturelm\b"],
    "liquid ai / insilico foundation model": [r"\bliquid[\s_-]?ai\b", r"insilico\s+medicine"],

    # === ADMET ===
    "admet-ai": [r"admet[_-]?ml", r"admet_ai", r"ADMET-AI"],
    "admetlab 3.0": [r"admetlab"],
    "chemprop": [r"\bchemprop\b"],
    "maplight + gnn": [r"\bmaplight\b"],
    "msformer-admet": [r"\bmsformer\b"],
    "mtan-admet": [r"\bmtan[_-]?admet\b", r"\bmtan\b"],
    "admet predictor 13": [r"admet[_\s-]?predictor"],

    # === PBPK / PK ===
    "ai-pbpk platform": [r"ai[_\s-]?pbpk", r"simcyp", r"certara"],
    "ml-pbpk / pbpk-ml": [r"pbpk[_\s-]?ml", r"ml[_\s-]?pbpk"],
    "deepliver activity": [r"\bdeepliver\b"],

    # === Toxicity ===
    "deeptox": [r"\bdeeptox\b"],
    "cardiotox": [r"\bcardiotox\b"],
    "bayesherg": [r"\bbayesh?erg\b"],
    "vitrobert": [r"\bvitrobert\b"],
    "hepatotoxicity portal (htp)": [r"\bhepatotoxicity[_\s-]?portal\b", r"\bHTP\b(?!\s*=)"],
    "neutoxpred 2.0": [r"\bneutoxpred\b"],

    # === CYP ===
    "cypreact": [r"\bcypreact\b"],
    "cyp-pro": [r"\bcyp[_\s-]?pro\b"],
    "mumcyp_net": [r"\bmumcyp\b"],

    # === Permeability / off-target ===
    "multitask learning (mtl) permeability": [r"mtl[_\s-]?permeability"],
    "ople (off-target)": [r"\bople\b"],

    # === Binding affinity ===
    "deepdta": [r"\bdeepdta\b", r"binding[_-]?ml"],
    "balm": [r"\bBALM\b"],
    "onionnet-2": [r"\bonionnet\b"],
    "ign": [r"\bIGN\b(?!\s*=)"],

    # === Target prediction ===
    "pidginv4": [r"\bpidgin\b"],
    "dtiam": [r"\bdtiam\b"],
    "chembl multitask": [r"chembl[_\s-]?multitask"],

    # === Synthetic accessibility ===
    "rascore": [r"\brascore\b"],
    "gasa": [r"\bGASA\b"],

    # === DDI ===
    "deepddi": [r"\bdeepddi\b"],
    "muffin": [r"\bMUFFIN\b"],

    # === Phase 1 / dose-response ===
    "aicmet": [r"\baicmet\b"],
    "gpt dose-response models": [r"gpt[_\s-]?dose[_\s-]?response"],

    # === Cross-layer / infrastructure ===
    # Note: deliberately do NOT alias "pk-sim" here. The codebase cites
    # PK-Sim's published physiological reference values throughout PBPK
    # code, but does not integrate the OSP Translational PKPD Framework
    # software. Counting those citations as integration would lie.
    "translational pkpd framework": [r"translational[_\s-]?pkpd"],
    "metadrp": [r"\bmetadrp\b"],
    "onnx runtime web": [r"onnxruntime[_\s-]?web", r"\bort-web\b", r"ort\.wasm"],
}


# Heuristic file-set classification: used to decide live vs scaffolded.
def is_backend_file(p: Path) -> bool:
    return AI_SERVICE_DIR in p.parents or p == AI_SERVICE_DIR / "main.py"


def is_frontend_file(p: Path) -> bool:
    return FRONTEND_SRC in p.parents


# Lines containing any of these substrings (case-insensitive) are
# treated as ASPIRATIONAL — i.e. they mention the model only as a
# future-work / wishlist / citation reference, not as an actual call.
ASPIRATIONAL_MARKERS: tuple[str, ...] = (
    "upgrade path",
    "would replace",
    "would be",
    "OOS today",
    "OOS —",
    "OOS -",
    "out of scope",
    "todo:",
    "fixme",
    "not yet",
    "cited reference",
    "replace this when",
    "could be",
    "future work",
    "wishlist",
    "aspirational",
    "inspired",
    "reference physiological",
    "reference value",
    "would swap",
    "swap in",
    "candidate model",
    "would call",
    "stub",
    "placeholder",
    "physiology reference",
    "replace with",
    "replacing the current",
    "would integrate",
    "to be wired",
    "not implemented",
    "documented in",
    "see also",
    "ref:",
)


_ASPIRATIONAL_MARKERS_LC: tuple[str, ...] = tuple(m.lower() for m in ASPIRATIONAL_MARKERS)


def is_aspirational_line(line: str) -> bool:
    s = line.lower()
    return any(marker in s for marker in _ASPIRATIONAL_MARKERS_LC)


@dataclass
class Hit:
    file: Path
    line_no: int
    line: str
    aspirational: bool = False


@dataclass
class ModelVerdict:
    name: str
    layer: str
    category: str
    sheet_status: str
    aliases: list[str]
    backend_hits: list[Hit] = field(default_factory=list)
    frontend_hits: list[Hit] = field(default_factory=list)

    @property
    def integration_state(self) -> str:
        real_backend = [h for h in self.backend_hits if not h.aspirational]
        real_frontend = [h for h in self.frontend_hits if not h.aspirational]
        if real_backend:
            return "live"
        if real_frontend:
            return "scaffolded"
        if self.backend_hits or self.frontend_hits:
            return "documented-only"
        if not self.aliases:
            return "unmapped"
        return "not-wired"

    @property
    def evidence_summary(self) -> str:
        n_b = len(self.backend_hits)
        n_f = len(self.frontend_hits)
        if n_b == 0 and n_f == 0:
            return "—"
        parts = []
        if n_b:
            parts.append(f"{n_b} backend ref(s)")
        if n_f:
            parts.append(f"{n_f} frontend ref(s)")
        return ", ".join(parts)


def iter_files() -> Iterable[Path]:
    if AI_SERVICE_DIR.exists():
        for p in AI_SERVICE_DIR.rglob("*.py"):
            if "__pycache__" in p.parts:
                continue
            yield p
    if FRONTEND_SRC.exists():
        for p in FRONTEND_SRC.rglob("*"):
            if p.suffix in {".ts", ".tsx", ".js", ".jsx"}:
                yield p


def load_files() -> list[tuple[Path, list[str]]]:
    out: list[tuple[Path, list[str]]] = []
    for p in iter_files():
        try:
            lines = p.read_text(errors="replace").splitlines()
        except Exception:
            continue
        out.append((p, lines))
    return out


def find_hits(patterns: list[str], corpus: list[tuple[Path, list[str]]]) -> tuple[list[Hit], list[Hit]]:
    backend: list[Hit] = []
    frontend: list[Hit] = []
    if not patterns:
        return backend, frontend
    compiled = [re.compile(pat, re.IGNORECASE) for pat in patterns]
    for path, lines in corpus:
        for i, line in enumerate(lines, start=1):
            for rx in compiled:
                if rx.search(line):
                    hit = Hit(
                        file=path,
                        line_no=i,
                        line=line.strip()[:160],
                        aspirational=is_aspirational_line(line),
                    )
                    if is_backend_file(path):
                        backend.append(hit)
                    elif is_frontend_file(path):
                        frontend.append(hit)
                    break
    return backend, frontend


def newest_xlsx() -> Path | None:
    candidates = sorted(
        glob.glob(str(REPO_ROOT / "attached_assets" / "HakaseAI_PreTrained_Models_Portfolio_*.xlsx"))
    )
    return Path(candidates[-1]) if candidates else None


def parse_master_portfolio(xlsx: Path) -> list[dict]:
    wb = load_workbook(xlsx, data_only=True)
    sheet_name = next((n for n in wb.sheetnames if n.lower().startswith("master")), None)
    if sheet_name is None:
        raise SystemExit(f"Master Portfolio sheet not found in {xlsx}")
    ws = wb[sheet_name]
    # Find header row by scanning for the cell containing "Model" + "Layer".
    header_row_idx = None
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        cells = [str(c).strip() if c is not None else "" for c in row]
        if "Model" in cells and "Layer" in cells:
            header_row_idx = i
            header = cells
            break
    if header_row_idx is None:
        raise SystemExit("Could not locate header row in Master Portfolio")
    rows: list[dict] = []
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        cells = [str(c).strip() if c is not None else "" for c in row]
        if not any(cells):
            continue
        record = {h: v for h, v in zip(header, cells) if h}
        if not record.get("Model"):
            continue
        rows.append(record)
    return rows


def build_verdicts(portfolio: list[dict], corpus: list[tuple[Path, list[str]]]) -> list[ModelVerdict]:
    verdicts: list[ModelVerdict] = []
    for row in portfolio:
        name = row.get("Model", "").strip()
        key = name.lower()
        aliases = ALIASES.get(key, [])
        backend, frontend = find_hits(aliases, corpus)
        verdicts.append(
            ModelVerdict(
                name=name,
                layer=row.get("Layer", ""),
                category=row.get("Category", ""),
                sheet_status=row.get("Status", ""),
                aliases=aliases,
                backend_hits=backend,
                frontend_hits=frontend,
            )
        )
    return verdicts


def write_csv(verdicts: list[ModelVerdict], out: Path) -> None:
    out.parent.mkdir(parents=True, exist_ok=True)
    with out.open("w", newline="") as f:
        w = csv.writer(f)
        w.writerow(
            ["Model", "Layer", "Category", "SheetStatus", "IntegrationState",
             "BackendRefs", "FrontendRefs", "EvidenceFiles"]
        )
        for v in verdicts:
            files = sorted({str(h.file.relative_to(REPO_ROOT))
                            for h in (v.backend_hits + v.frontend_hits)})
            w.writerow([
                v.name, v.layer, v.category, v.sheet_status,
                v.integration_state, len(v.backend_hits), len(v.frontend_hits),
                "; ".join(files),
            ])


def write_markdown(verdicts: list[ModelVerdict], out: Path, xlsx: Path) -> None:
    out.parent.mkdir(parents=True, exist_ok=True)
    by_state = {"live": [], "scaffolded": [], "documented-only": [],
                "not-wired": [], "unmapped": []}
    for v in verdicts:
        by_state[v.integration_state].append(v)

    total = len(verdicts)
    n_live = len(by_state["live"])
    n_scaf = len(by_state["scaffolded"])
    n_doc = len(by_state["documented-only"])
    n_nope = len(by_state["not-wired"])
    n_unmapped = len(by_state["unmapped"])

    lines: list[str] = []
    lines.append("# Model Portfolio — Integration Audit")
    lines.append("")
    lines.append(f"_Generated by `scripts/audit_model_portfolio.py` against `{xlsx.relative_to(REPO_ROOT)}`._")
    lines.append("")
    lines.append("This file is **auto-generated**. Do not edit by hand. Re-run the script after any backend / frontend integration change.")
    lines.append("")
    lines.append("## Summary")
    lines.append("")
    lines.append(f"- **Catalogued in spreadsheet**: {total}")
    lines.append(f"- **Live** (backend endpoint or backend module references the model): {n_live}")
    lines.append(f"- **Scaffolded** (referenced only in frontend code, no backend): {n_scaf}")
    lines.append(f"- **Documented-only** (named only in comments / wishlist text): {n_doc}")
    lines.append(f"- **Not wired** (no code reference at all): {n_nope}")
    if n_unmapped:
        lines.append(f"- **Unmapped** (no alias in the auditor — add one to `scripts/audit_model_portfolio.py`): {n_unmapped}")
    lines.append("")
    lines.append("## Verdict by model")
    lines.append("")
    lines.append("| Model | Layer | Sheet status | Integration state | Evidence |")
    lines.append("| --- | --- | --- | --- | --- |")
    for v in verdicts:
        badge = {
            "live": "✅ live",
            "scaffolded": "🟡 scaffolded",
            "documented-only": "📝 doc-only",
            "not-wired": "⬜ not wired",
            "unmapped": "❓ unmapped",
        }[v.integration_state]
        lines.append(
            f"| {v.name} | {v.layer} | {v.sheet_status} | {badge} | {v.evidence_summary} |"
        )
    lines.append("")
    for state, label in [("live", "Live integrations"),
                         ("scaffolded", "Scaffolded (frontend only)"),
                         ("documented-only", "Documented-only (wishlist / citation)"),
                         ("unmapped", "Unmapped models")]:
        bucket = by_state[state]
        if not bucket:
            continue
        lines.append(f"## {label}")
        lines.append("")
        for v in bucket:
            lines.append(f"### {v.name}")
            lines.append("")
            lines.append(f"- Layer: {v.layer}")
            lines.append(f"- Category: {v.category}")
            lines.append(f"- Sheet status: {v.sheet_status}")
            lines.append(f"- Integration state: **{v.integration_state}**")
            if v.aliases:
                lines.append(f"- Aliases searched: `{', '.join(v.aliases)}`")
            else:
                lines.append("- Aliases searched: _(none — add to ALIASES)_")
            evidence = (v.backend_hits + v.frontend_hits)[:6]
            if evidence:
                lines.append("- Top evidence:")
                for h in evidence:
                    rel = h.file.relative_to(REPO_ROOT)
                    lines.append(f"  - `{rel}:{h.line_no}` — `{h.line}`")
            lines.append("")

    out.write_text("\n".join(lines))


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--xlsx", type=Path, default=None,
                        help="Path to the portfolio spreadsheet (defaults to newest in attached_assets/)")
    args = parser.parse_args()

    xlsx = args.xlsx or newest_xlsx()
    if xlsx is None or not xlsx.exists():
        sys.stderr.write("No portfolio xlsx found. Pass --xlsx <path>.\n")
        return 2

    print(f"[audit] reading {xlsx.relative_to(REPO_ROOT)}")
    portfolio = parse_master_portfolio(xlsx)
    print(f"[audit] {len(portfolio)} models in spreadsheet")

    print("[audit] scanning codebase…")
    corpus = load_files()
    print(f"[audit] {len(corpus)} source files scanned")

    verdicts = build_verdicts(portfolio, corpus)

    md_path = DOCS_DIR / "model-portfolio-audit.md"
    csv_path = DOCS_DIR / "model-portfolio-audit.csv"
    write_markdown(verdicts, md_path, xlsx)
    write_csv(verdicts, csv_path)

    counts = {s: 0 for s in ("live", "scaffolded", "documented-only",
                              "not-wired", "unmapped")}
    for v in verdicts:
        counts[v.integration_state] += 1

    print()
    print("=== Audit summary ===")
    for k, v in counts.items():
        print(f"  {k:>11}: {v}")
    print(f"\nWrote: {md_path.relative_to(REPO_ROOT)}")
    print(f"Wrote: {csv_path.relative_to(REPO_ROOT)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
